from openpyxl import Workbook
from typing import Dict, Any, List


# ---------------------------
# CONFIG (editável sem saber programar)
# ---------------------------

SHEETS: Dict[str, Dict[str, Any]] = {
    "Produtos": {
        "columns": [
            "ID_PRODUTO",
            "NOME",
            "CATEGORIA",
            "CUSTO_UNITARIO",
            "PRECO_VENDA",
            "ATIVO",
        ],
        "rows": [
            [1, "Cheeseburger Artesanal", "Lanche", 8.50, 25.00, "S"],
            [2, "Batata Frita Média", "Acompanhamento", 3.00, 12.00, "S"],
            [3, "Refrigerante Lata 350ml", "Bebida", 2.80, 8.00, "S"],
            [4, "Combo Burger + Batata + Refri", "Combo", 14.30, 38.00, "S"],
            [5, "Água Mineral 500ml", "Bebida", 1.50, 5.00, "S"],
            [6, "Milkshake Chocolate 400ml", "Sobremesa", 4.20, 18.00, "S"],
        ],
    },
    "Estoque_Inicial": {
        "columns": [
            "ID_PRODUTO",
            "NOME",
            "QTD_ESTOQUE_INICIAL",
            "UNIDADE",
        ],
        "rows": [
            [1, "Cheeseburger Artesanal", 50, "un"],
            [2, "Batata Frita Média", 80, "porção"],
            [3, "Refrigerante Lata 350ml", 120, "lata"],
            [4, "Combo Burger + Batata + Refri", 40, "combo"],
            [5, "Água Mineral 500ml", 60, "garrafa"],
            [6, "Milkshake Chocolate 400ml", 35, "copo"],
        ],
    },
    "Vendas_Itens": {
        "columns": [
            "DATA",
            "HORA",
            "ID_VENDA",
            "ID_PRODUTO",
            "PRODUTO",
            "QTD",
            "VALOR_UNITARIO",
            "TOTAL_LINHA",
            "FORMA_PAGTO",
        ],
        "rows": [
            ["2025-10-23", "12:15", 1001, 4, "Combo Burger + Batata + Refri", 2, 38.00, 76.00, "Cartão"],
            ["2025-10-23", "12:15", 1001, 6, "Milkshake Chocolate 400ml", 1, 18.00, 18.00, "Cartão"],
            ["2025-10-23", "13:02", 1002, 1, "Cheeseburger Artesanal", 1, 25.00, 25.00, "Pix"],
            ["2025-10-23", "13:02", 1002, 2, "Batata Frita Média", 1, 12.00, 12.00, "Pix"],
            ["2025-10-23", "13:02", 1002, 3, "Refrigerante Lata 350ml", 1, 8.00, 8.00, "Pix"],
            ["2025-10-23", "14:37", 1003, 5, "Água Mineral 500ml", 2, 5.00, 10.00, "Dinheiro"],
            ["2025-10-22", "20:41", 1004, 6, "Milkshake Chocolate 400ml", 1, 18.00, 18.00, "Cartão"],
            ["2025-10-22", "20:41", 1004, 1, "Cheeseburger Artesanal", 1, 25.00, 25.00, "Cartão"],
        ],
    },
    "Fluxo_Caixa": {
        "columns": [
            "DATA",
            "TIPO",
            "DESCRICAO",
            "ENTRADA",
            "SAIDA",
            "SALDO_ACUMULADO",
        ],
        "rows": [
            ["2025-10-23", "Saldo Inicial", "Abertura do Caixa", 500.00, 0.00, 500.00],
            ["2025-10-23", "Venda 1001", "Cartão", 76.00, 0.00, 576.00],
            ["2025-10-23", "Venda 1002", "Pix", 45.00, 0.00, 621.00],
            ["2025-10-23", "Venda 1003", "Dinheiro", 10.00, 0.00, 631.00],
            ["2025-10-23", "Despesa Insumos", "Compra pão/queijo/carne", 0.00, 120.00, 511.00],
            ["2025-10-23", "Retirada Dono", "Pró-labore diário", 0.00, 50.00, 461.00],
        ],
    },
}


# ---------------------------
# FUNÇÕES DE NEGÓCIO
# ---------------------------

def validar_dados(planilhas: Dict[str, Dict[str, Any]]) -> List[str]:
    """
    Garante que todas as linhas tenham o mesmo número de colunas que o cabeçalho.
    Retorna lista de erros. Lista vazia = tudo certo.
    """
    erros = []
    for nome_sheet, cfg in planilhas.items():
        num_cols = len(cfg["columns"])
        for i, row in enumerate(cfg["rows"], start=1):
            if len(row) != num_cols:
                erros.append(
                    f"Aba '{nome_sheet}': linha {i} tem {len(row)} colunas, mas o cabeçalho tem {num_cols}."
                )
    return erros


def gerar_workbook(planilhas: Dict[str, Dict[str, Any]]) -> Workbook:
    """
    Cria o Workbook (arquivo Excel em memória) com todas as abas e dados.
    """
    wb = Workbook()
    primeira = True

    for nome_sheet, cfg in planilhas.items():
        if primeira:
            ws = wb.active
            ws.title = nome_sheet
            primeira = False
        else:
            ws = wb.create_sheet(nome_sheet)

        # Cabeçalho
        ws.append(cfg["columns"])

        # Linhas
        for row in cfg["rows"]:
            ws.append(row)

    return wb


def salvar_excel(wb: Workbook, caminho_arquivo: str) -> str:
    """
    Salva o workbook no disco e retorna o caminho.
    """
    wb.save(caminho_arquivo)
    return caminho_arquivo


def testar_geracao() -> dict:
    """
    Executa uma bateria de testes simples de integridade.
    - valida estrutura
    - gera workbook
    - confere abas
    - confere cabeçalhos
    - salva arquivo
    Retorna um resumo (pode ser usado como log).
    """
    # 1. validar consistência
    erros = validar_dados(SHEETS)
    assert not erros, f"Falha na validação de dados: {erros}"

    # 2. gerar workbook
    wb = gerar_workbook(SHEETS)

    # 3. validar abas
    abas_esperadas = set(SHEETS.keys())
    abas_geradas = set(wb.sheetnames)
    assert abas_esperadas == abas_geradas, f"Abas não batem. Esperado {abas_esperadas}, Gerado {abas_geradas}"

    # 4. validar cabeçalhos
    for nome_sheet, cfg in SHEETS.items():
        ws = wb[nome_sheet]
        header_lido = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        assert header_lido == cfg["columns"], f"Cabeçalho diferente na aba {nome_sheet}"

    # 5. salvar
    caminho = "lanchonete.xlsx"
    salvar_excel(wb, caminho)

    return {
        "ok": True,
        "arquivo": caminho,
        "abas": wb.sheetnames,
        "linhas_por_aba": {
            nome: len(list(wb[nome].iter_rows()))
            for nome in wb.sheetnames
        }
    }


if __name__ == "_main_":
    resultado = testar_geracao()
    print("Planilha gerada com sucesso!")
    print("Resumo:", resultado)