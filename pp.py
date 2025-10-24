#!/usr/bin/env python3
# -- coding: utf-8 --

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: A biblioteca 'openpyxl' não está instalada!")
    print("Execute: pip install openpyxl")
    exit(1)

from pathlib import Path


class ExcelGenerator:
    """Classe para gerar planilhas Excel de forma limpa e organizada."""
    
    def _init_(self, filename: str):
        self.filename = filename
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        
    def set_sheet_name(self, name: str) -> None:
        """Define o nome da planilha."""
        self.sheet.title = name
        
    def add_header(self, headers: list, row: int = 1) -> None:
        """Adiciona cabeçalhos formatados à planilha."""
        for col, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, size=12, color='FFFFFF')
    
    def add_data(self, data: list[list], start_row: int = 2) -> None:
        """Adiciona dados à planilha."""
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=1):
                self.sheet.cell(row=row_idx, column=col_idx, value=value)
    
    def auto_adjust_columns(self) -> None:
        """Ajusta automaticamente a largura das colunas."""
        for column in self.sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)
            self.sheet.column_dimensions[column_letter].width = adjusted_width
    
    def save(self) -> None:
        """Salva a planilha no arquivo especificado."""
        output_path = Path(self.filename)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.filename)
        print(f"✓ Planilha salva em: {self.filename}")


def create_sample_spreadsheet() -> None:
    """Exemplo de uso: cria uma planilha de exemplo."""
    
    # Dados de exemplo
    headers = ['ID', 'Nome', 'Email', 'Departamento', 'Salário']
    data = [
        [1, 'João Silva', 'joao@email.com', 'TI', 5500.00],
        [2, 'Maria Santos', 'maria@email.com', 'RH', 4800.00],
        [3, 'Pedro Costa', 'pedro@email.com', 'Vendas', 6200.00],
        [4, 'Ana Oliveira', 'ana@email.com', 'Marketing', 5000.00],
        [5, 'Carlos Souza', 'carlos@email.com', 'TI', 5800.00],
    ]
    
    # Gerar planilha
    excel = ExcelGenerator('funcionarios.xlsx')
    excel.set_sheet_name('Funcionários')
    excel.add_header(headers)
    excel.add_data(data)
    excel.auto_adjust_columns()
    excel.save()


if __name__ == "_main_":
    try:
        print("Iniciando geração da planilha...")
        create_sample_spreadsheet()
        print("\n✓ Processo concluído com sucesso!")
    except Exception as e:
        print(f"\n✗ Erro ao gerar planilha: {e}")
        import traceback
        traceback.print_exc()